import streamlit as st
import pandas as pd
import os
from xlsx2html import xlsx2html
import streamlit.components.v1 as components
from langchain_experimental.agents.agent_toolkits import create_pandas_dataframe_agent
from langchain_google_genai import ChatGoogleGenerativeAI
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
import io
from streamlit_autorefresh import st_autorefresh
# 1. Configuración de la página (Wide mode)
st.set_page_config(page_title="Payroll", layout="wide", initial_sidebar_state="collapsed")

st.markdown("""
<style>
    /* Ocultar elementos de Streamlit por defecto */
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    header {visibility: hidden;}

    /* Ocultar padding superior para maximizar espacio */
    .block-container {
        padding-top: 1rem !important;
        padding-bottom: 0rem !important;
        padding-left: 1rem !important;
        padding-right: 1rem !important;
        max-width: 100% !important;
    }

    /* ==== CONTORNO DEL PANEL LATERAL (TODA LA COLUMNA DERECHA) ==== */
    [data-testid="stColumn"]:has(#right-panel-marker),
    [data-testid="column"]:has(#right-panel-marker) {
        border: 2px solid rgba(74, 144, 226, 0.4) !important;
        border-radius: 20px !important;
        padding: 15px 15px 15px 15px !important;
        background-color: rgba(245, 248, 252, 0.9) !important;
        box-shadow: -5px 0 25px rgba(0, 0, 0, 0.05) !important;
        height: 100% !important;
    }
    
    @media (prefers-color-scheme: dark) {
        [data-testid="stColumn"]:has(#right-panel-marker),
        [data-testid="column"]:has(#right-panel-marker) {
            background-color: rgba(15, 20, 25, 0.7) !important;
            border: 2px solid rgba(74, 144, 226, 0.3) !important;
            box-shadow: -5px 0 25px rgba(0, 0, 0, 0.2) !important;
        }
    }

    /* ==== DISEÑO PREMIUM DEL CUADRO DE INTERACCIÓN (CHAT) ==== */

    /* Diseño de las burbujas de chat */
    [data-testid="stChatMessage"] {
        background: rgba(30, 34, 45, 0.5);
        border: 1px solid rgba(255, 255, 255, 0.08);
        border-radius: 18px;
        padding: 12px 18px;
        margin-bottom: 16px;
        box-shadow: 0 8px 24px rgba(0, 0, 0, 0.12);
        backdrop-filter: blur(12px);
        -webkit-backdrop-filter: blur(12px);
        transition: transform 0.2s ease, border 0.2s ease;
        line-height: 1.5;
        font-family: 'Inter', -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
    }

    [data-testid="stChatMessage"]:hover {
        transform: translateY(-2px);
        border: 1px solid rgba(255, 255, 255, 0.15);
        background: rgba(30, 34, 45, 0.7);
    }

    /* Diferenciar color de texto general para nitidez */
    [data-testid="stChatMessage"] p {
        color: #E2E8F0 !important;
        font-size: 0.95rem;
    }

    /* Caja de texto del input flotante (Chat Input) */
    .stChatFloatingInputContainer {
        background: transparent !important;
        padding-bottom: 25px !important;
    }

    .stChatFloatingInputContainer > div {
        background: linear-gradient(135deg, rgba(25, 28, 36, 0.9) 0%, rgba(18, 20, 26, 0.95) 100%) !important;
        border: 1px solid rgba(255, 255, 255, 0.1) !important;
        border-radius: 24px !important;
        box-shadow: 0 10px 30px rgba(0, 0, 0, 0.4), inset 0 1px 1px rgba(255,255,255,0.05) !important;
        backdrop-filter: blur(20px) !important;
        transition: all 0.3s ease !important;
    }

    /* Brillo al enfocar el cuadro de texto */
    .stChatFloatingInputContainer > div:focus-within {
        border: 1px solid rgba(74, 144, 226, 0.6) !important;
        box-shadow: 0 0 20px rgba(74, 144, 226, 0.3), 0 10px 30px rgba(0,0,0,0.4) !important;
        background: rgba(20, 22, 28, 0.95) !important;
    }

    /* Contorno especial para los mensajes de la IA (Asistente) */
    [data-testid="stChatMessage"]:has([data-testid*="AvatarAssistant"]), 
    [data-testid="stChatMessage"]:has(svg),
    [data-testid="stChatMessage"]:nth-child(even) {
        /* Usamos nth-child(even) o has(svg) como fallback en caso de que cambien los selectores internos de Streamlit */
    }
    
    [data-testid="stChatMessage"]:has([data-testid*="assistant"]),
    [data-testid="stChatMessage"]:has(svg:not([data-testid*="user"])) {
        border: 1.5px solid rgba(74, 144, 226, 0.7) !important;
        box-shadow: 0 0 15px rgba(74, 144, 226, 0.2) !important;
        background: rgba(25, 35, 50, 0.6) !important;
    }

    /* Avatares */
    [data-testid="stChatMessageAvatar"], [data-testid*="Avatar"] {
        border-radius: 12px !important;
        box-shadow: 0 4px 10px rgba(0,0,0,0.3) !important;
        border: 1px solid rgba(255,255,255,0.1);
    }
    
    /* Contorno extra al avatar de la IA */
    [data-testid*="assistant"] [data-testid*="Avatar"],
    [data-testid="stChatMessage"]:has([data-testid*="assistant"]) [data-testid="stChatMessageAvatar"] {
        border: 2px solid rgba(74, 144, 226, 0.9) !important;
        box-shadow: 0 0 12px rgba(74, 144, 226, 0.4) !important;
    }

    /* Ajuste para los iconos dentro del input */
    [data-testid="stChatInputSubmitButton"] {
        color: #4a90e2 !important;
        transition: color 0.2s ease, transform 0.2s ease;
    }
    
    [data-testid="stChatInputSubmitButton"]:hover {
        color: #6fb0ff !important;
        transform: scale(1.1);
    }
</style>
""", unsafe_allow_html=True)

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass
EXCEL_FILE = "PAYROLL.xlsx"
HTML_FILE = "temp_excel_view.html"

def sync_google_drive():
    try:
        SCOPES = ['https://www.googleapis.com/auth/drive.readonly']
        SERVICE_ACCOUNT_FILE = 'credentials.json'
        if not os.path.exists(SERVICE_ACCOUNT_FILE):
            st.sidebar.error("Falta credentials.json para conectar a Google Drive.")
            return False
        
        creds = service_account.Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)
        service = build('drive', 'v3', credentials=creds)
        
        file_id = '19igCbEX2mtp8GRj14QDmRXYMVM7ecdgz'
        request = service.files().get_media(fileId=file_id)
        fh = io.BytesIO()
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        with st.sidebar.status("Descargando desde Google Drive...", expanded=True) as status:
            while done is False:
                status_chunk, done = downloader.next_chunk()
            fh.seek(0)
            with open(EXCEL_FILE, 'wb') as f:
                f.write(fh.read())
            
            # Guardar la fecha exacta de esta versión para comparaciones futuras
            try:
                current_cloud_time = service.files().get(fileId=file_id, fields="modifiedTime").execute().get('modifiedTime')
                if current_cloud_time:
                    with open('last_sync.txt', 'w') as f:
                        f.write(current_cloud_time)
            except:
                pass

            if os.path.exists(HTML_FILE):
                os.remove(HTML_FILE)
            status.update(label="¡Sincronizado correctamente!", state="complete", expanded=False)
        return True
    except Exception as e:
        st.sidebar.error(f"Error conectando a Google Drive: {e}")
        return False

with st.sidebar:
    st.markdown("### ☁️ Sincronización Nube")
    if st.button("🔄 Actualizar Datos desde Google Sheets"):
        if sync_google_drive():
            st.rerun()

    st.divider()
    st.markdown("### 🔑 Configuración IA")
    saved_key = os.environ.get("GEMINI_API_KEY", "")
    new_key = st.text_input("Gemini API Key", value=saved_key, type="password", placeholder="Pega tu clave aquí")
    
    if new_key and new_key != saved_key:
        env_path = ".env"
        lines = []
        if os.path.exists(env_path):
            with open(env_path, "r", encoding="utf-8") as f:
                lines = f.readlines()
        with open(env_path, "w", encoding="utf-8") as f:
            found = False
            for line in lines:
                if line.startswith("GEMINI_API_KEY="):
                    f.write(f"GEMINI_API_KEY={new_key}\n")
                    found = True
                else:
                    f.write(line)
            if not found:
                f.write(f"GEMINI_API_KEY={new_key}\n")
        os.environ["GEMINI_API_KEY"] = new_key
        st.rerun()
        
    if st.button("Limpiar Key"):
        env_path = ".env"
        if os.path.exists(env_path):
            with open(env_path, "r", encoding="utf-8") as f:
                lines = f.readlines()
            with open(env_path, "w", encoding="utf-8") as f:
                for line in lines:
                    if not line.startswith("GEMINI_API_KEY="):
                        f.write(line)
        if "GEMINI_API_KEY" in os.environ:
            del os.environ["GEMINI_API_KEY"]
        st.rerun()

# ----------------- AUTO-SYNC EN TIEMPO REAL -----------------
# 1. Mantener la aplicación despierta y refrescando cada 60 segundos
st_autorefresh(interval=60 * 1000, key="data_autorefresh")

# 2. Función súper rápida para revisar si hubo cambios en Drive (caché de 60s)
@st.cache_data(ttl=60)
def get_cloud_modified_time():
    try:
        SCOPES = ['https://www.googleapis.com/auth/drive.readonly']
        SERVICE_ACCOUNT_FILE = 'credentials.json'
        if not os.path.exists(SERVICE_ACCOUNT_FILE): return None
        creds = service_account.Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)
        service = build('drive', 'v3', credentials=creds)
        file_id = '19igCbEX2mtp8GRj14QDmRXYMVM7ecdgz'
        metadata = service.files().get(fileId=file_id, fields="modifiedTime").execute()
        return metadata.get('modifiedTime')
    except Exception:
        return None

# 3. Lógica de auto-descarga si detecta que la versión en la nube es más nueva
cloud_time = get_cloud_modified_time()
if cloud_time:
    last_sync_time = None
    if os.path.exists('last_sync.txt'):
        with open('last_sync.txt', 'r') as f:
            last_sync_time = f.read().strip()
            
    if last_sync_time != cloud_time:
        # Solo mostramos el toast si el archivo ya existía (es una actualización, no la primera descarga)
        if os.path.exists(EXCEL_FILE):
            st.toast("Actualización detectada en la nube. Sincronizando...", icon="🔄")
        if sync_google_drive():
            st.rerun()
# -----------------------------------------------------------

if not os.path.exists(EXCEL_FILE):
    st.info("Sincronizando información desde Google Drive por primera vez...")
    if not sync_google_drive():
        st.error(f"No se pudo descargar de Google Drive y el archivo {EXCEL_FILE} no existe localmente.")
        st.stop()
    else:
        st.rerun()

# Dividir en dos columnas (85% Excel, 15% Chat)
col1, col2 = st.columns([5.5, 1])

with col1:
    # Revisar si se necesita regenerar el HTML
    regenerate = False
    if not os.path.exists(HTML_FILE):
        regenerate = True
    else:
        # Regenerar si el archivo Excel es más nuevo que el HTML
        if os.path.getmtime(EXCEL_FILE) > os.path.getmtime(HTML_FILE):
            regenerate = True

    if regenerate:
        try:
            with st.spinner("Procesando estilos originales..."):
                import openpyxl
                
                temp_excel = "temp_clean.xlsx"
                wb = openpyxl.load_workbook(EXCEL_FILE)
                ws = wb.active
                
                # Encontrar dinámicamente la última columna con encabezado o datos (revisando las primeras 15 filas)
                max_c = ws.max_column
                last_valid_col = 1
                for c in range(1, max_c + 1):
                    if any(ws.cell(row=r, column=c).value is not None for r in range(1, min(15, ws.max_row + 1))):
                        last_valid_col = c

                # Eliminar columnas vacías después de la última válida
                if ws.max_column > last_valid_col:
                    ws.delete_cols(last_valid_col + 1, ws.max_column - last_valid_col)
                    
                # Buscar la última fila con datos reales
                max_r = ws.max_row
                last_valid_row = 1
                for r in range(1, max_r + 1):
                    if any(ws.cell(row=r, column=c).value is not None for c in range(1, last_valid_col + 1)):
                        last_valid_row = r
                
                if max_r > last_valid_row:
                    ws.delete_rows(last_valid_row + 1, max_r - last_valid_row)
                    
                wb.save(temp_excel)
                
                xlsx2html(temp_excel, HTML_FILE)
                
                if os.path.exists(temp_excel):
                    os.remove(temp_excel)
                
                # Post-procesar HTML para inyectar CSS interno del iframe
                with open(HTML_FILE, "r", encoding="utf-8") as f:
                    html_content = f.read()
                
                # Inyectar estilos para que el excel no tenga márgenes y se adapte al dark mode
                custom_iframe_css = """
                <style>
                    body {
                        margin: 0;
                        padding: 0;
                        background-color: #ffffff;
                    }
                    table {
                        border-collapse: collapse !important;
                        min-width: 100% !important;
                    }
                    col {
                        width: auto !important;
                    }
                    td, th {
                        border: none !important;
                        padding: 12px 18px !important;
                        font-size: 15px !important;
                        line-height: 1.4 !important;
                        min-width: 150px;
                    }
                    
                    /* Si la fila tiene un cuadro gris (varios tonos), pintar TODA la fila uniformemente de gris claro */
                    tr:has(td[style*="BFBFBF"]) td,
                    tr:has(td[style*="808080"]) td,
                    tr:has(td[style*="7F7F7F"]) td,
                    tr:has(td[style*="D8D8D8"]) td,
                    tr:has(td[style*="bfbfbf"]) td {
                        background-color: #f2f4f7 !important;
                    }
                    
                    /* Borde inferior sutil para las filas de datos */
                    tr:nth-child(n+4) td {
                        border-bottom: 1px solid #eef2f6 !important;
                    }
                </style>
                """
                html_content = html_content.replace('</head>', f'{custom_iframe_css}</head>')
                
                # Arreglar el color del texto en las celdas con fondo azul oscuro para que sea blanco y visible
                html_content = html_content.replace('background-color: #002060', 'background-color: #002060; color: #ffffff !important')
                html_content = html_content.replace('color: #000000;font-size: 11.0px;height: 23.0pt', 'color: #ffffff !important;font-size: 15.0px;height: 23.0pt')
                
                # Inyectar Javascript para filtros
                filter_script = """
                <script>
                document.addEventListener("DOMContentLoaded", function() {
                    // Buscar la fila de encabezados
                    var allTds = document.querySelectorAll("td, th");
                    var headerRow = null;
                    for (var i = 0; i < allTds.length; i++) {
                        if (allTds[i].innerText.includes("EMPLOYEE NAME")) {
                            headerRow = allTds[i].parentElement;
                            break;
                        }
                    }
                    
                    if (headerRow) {
                        var headers = headerRow.children;
                        var tableRows = headerRow.parentElement.children;
                        var headerIndex = Array.prototype.indexOf.call(tableRows, headerRow);
                        
                        for (let i = 0; i < headers.length; i++) {
                            // Ignorar celdas vacías
                            if (headers[i].innerText.trim() === "") continue;
                            
                            let select = document.createElement("select");
                            select.innerHTML = '<option value="">Filtrar...</option>';
                            select.style.display = 'block';
                            select.style.marginTop = '8px';
                            select.style.padding = '6px';
                            select.style.borderRadius = '6px';
                            select.style.border = '1px solid rgba(74, 144, 226, 0.5)';
                            select.style.backgroundColor = '#ffffff';
                            select.style.color = '#333333';
                            select.style.width = '100%';
                            select.style.fontSize = '13px';
                            select.style.cursor = 'pointer';
                            select.style.boxShadow = '0 2px 5px rgba(0,0,0,0.05)';
                            
                            let uniqueVals = new Set();
                            for (let j = headerIndex + 1; j < tableRows.length; j++) {
                                let cell = tableRows[j].children[i];
                                if (cell && cell.innerText.trim() !== "") {
                                    uniqueVals.add(cell.innerText.trim());
                                }
                            }
                            
                            let sortedVals = Array.from(uniqueVals).sort();
                            sortedVals.forEach(val => {
                                let option = document.createElement("option");
                                option.value = val;
                                option.text = val.length > 25 ? val.substring(0, 25) + '...' : val;
                                select.appendChild(option);
                            });
                            
                            select.addEventListener("change", function() {
                                let allSelects = headerRow.querySelectorAll("select");
                                for (let j = headerIndex + 1; j < tableRows.length; j++) {
                                    let showRow = true;
                                    for (let k = 0; k < allSelects.length; k++) {
                                        if (allSelects[k].value !== "") {
                                            let cell = tableRows[j].children[k];
                                            if (!cell || cell.innerText.trim() !== allSelects[k].value) {
                                                showRow = false;
                                                break;
                                            }
                                        }
                                    }
                                    tableRows[j].style.display = showRow ? "" : "none";
                                }
                            });
                            
                            headers[i].appendChild(select);
                        }
                    }
                });
                </script>
                """
                if '</body>' in html_content:
                    html_content = html_content.replace('</body>', f'{filter_script}</body>')
                else:
                    html_content += filter_script
                
                with open(HTML_FILE, "w", encoding="utf-8") as f:
                    f.write(html_content)
                    
        except Exception as e:
            st.error(f"Error renderizando Excel a HTML: {e}")

    # Mostrar el HTML
    if os.path.exists(HTML_FILE):
        with open(HTML_FILE, "r", encoding="utf-8") as f:
            html_data = f.read()
        
        # Contenedor con altura fija y scroll
        components.html(html_data, height=950, scrolling=True)


with col2:
    st.markdown("<span id='right-panel-marker'></span>", unsafe_allow_html=True)

    api_key = os.environ.get("GEMINI_API_KEY", "")

    # Inicializar estado
    if "messages" not in st.session_state:
        st.session_state.messages = []

    # Contenedor scrolleable para el chat
    # Ajustamos la altura a 760px para compensar el espacio que ocupa el input de la API Key arriba
    # y así mantener la parte inferior alineada con la tabla de 850px.
    chat_container = st.container(height=760, border=False)
    
    with chat_container:
        for msg in st.session_state.messages:
            with st.chat_message(msg["role"]):
                st.markdown(msg["content"])

    if not api_key:
        st.warning("⚠️ Por favor ingresa tu API Key arriba.")
        st.chat_input("Ingresa tu API Key para chatear...", disabled=True)
    else:
        if prompt := st.chat_input("Ej: ¿Cuánto se le paga al equipo estacional?"):
            
            st.session_state.messages.append({"role": "user", "content": prompt})
            with chat_container:
                with st.chat_message("user"):
                    st.markdown(prompt)

                with st.chat_message("assistant"):
                    with st.spinner("Analizando información..."):
                        try:
                            df = pd.read_excel(EXCEL_FILE)
                            context_df = df.to_markdown()
                            llm = ChatGoogleGenerativeAI(
                                model="gemini-2.5-flash", 
                                google_api_key=api_key, 
                                temperature=0.2
                            )
                            
                            from langchain_core.messages import SystemMessage, HumanMessage, AIMessage
                            
                            # Construir el historial de mensajes
                            langchain_messages = [
                                SystemMessage(content=f"Eres un asistente experto en analizar nóminas. Responde cualquier pregunta que te haga el usuario basándote únicamente en el siguiente documento de nómina (PAYROLL):\n\n{context_df}")
                            ]
                            
                            # Añadir historial del chat (excluyendo el prompt actual que ya se añadió en st.session_state)
                            for msg in st.session_state.messages:
                                if msg["role"] == "user":
                                    langchain_messages.append(HumanMessage(content=msg["content"]))
                                elif msg["role"] == "assistant":
                                    langchain_messages.append(AIMessage(content=msg["content"]))
                            
                            # Generar la respuesta
                            response = llm.invoke(langchain_messages)
                            answer = response.content
                            
                            st.markdown(answer)
                            st.session_state.messages.append({"role": "assistant", "content": answer})
                        except Exception as e:
                            import requests
                            error_msg = f"Error: {e}"
                            try:
                                url = f"https://generativelanguage.googleapis.com/v1beta/models?key={api_key}"
                                res = requests.get(url)
                                models_data = res.json()
                                available = [m['name'].replace('models/', '') for m in models_data.get('models', []) if 'generateContent' in m.get('supportedGenerationMethods', [])]
                                error_msg += f"\n\nModelos disponibles en tu cuenta: {', '.join(available)}"
                            except Exception as ex:
                                error_msg += f"\n\n(No se pudieron cargar los modelos disponibles: {ex})"
                            st.error(error_msg)
                            st.session_state.messages.append({"role": "assistant", "content": error_msg})
